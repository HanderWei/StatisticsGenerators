from openpyxl import load_workbook
import re

key_value_p = re.compile('\w*[:]+\s*\w*')


# def parse_file(file_name, table_name):
#     """解析文件, 返回数据字典列表"""
#     wb = load_workbook(filename=file_name)
#     ws = wb[table_name]
#     result_list = []
#     for row in ws.rows:
#         if should_parse_this_row(row):
#             result_list.append(parse_row(row))
#     return result_list


def parse_file(file_name, table_name, start=0, end=0):
    """解析文件, 返回数据字典列表"""
    wb = load_workbook(filename=file_name)
    ws = wb[table_name]
    result_list = []
    if start == 0 and end == 0:
        rows = ws.rows
    else:
        rows = ws.iter_rows(min_row=start, max_row=end)
    for row in rows:
        if should_parse_this_row(row):
            result_list.append(parse_row(row))
    return result_list


def should_parse_this_row(row):
    """判断当前row是否符合解析的要求"""
    if row[7].value is not None and str(row[7].value).strip()[0] == '{' and str(row[7].value).strip()[-1] == '}':
        return True
    else:
        return False


def parse_row(row):
    """解析列"""
    return {'id': str(row[0].value).strip(), 'version': str(row[1].value).strip(),
            'condition': str(row[2].value).strip(), 'category': str(row[3].value).strip(),
            'eventId': str(row[4].value).strip(), 'label': str(row[5].value).strip(),
            'actionId': str(row[6].value).strip(), 'data': get_key_values(str(row[7].value).strip())}


def get_key_values(kv_str):
    """解析表格中最后一项"""
    kv_str = kv_str.lstrip('{').rstrip('}')
    kws_list = key_value_p.findall(kv_str)
    kwargs = {}
    for kw_str in kws_list:
        words = kw_str.split(':')
        key = words[0].strip()
        value = words[-1].strip()
        kwargs[key] = value
    return kwargs
